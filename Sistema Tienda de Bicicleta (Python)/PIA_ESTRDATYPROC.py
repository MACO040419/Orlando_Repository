import csv
import openpyxl
import datetime
import pandas as pd
import sqlite3
from sqlite3 import Error
import matplotlib.pyplot as plt

ruta_actual = ['Menú Principal']
rentas_por_cliente = {}
unidades = {}
colores = [
    "ROJO",
    "VERDE",
    "AZUL",
    "AMARILLO",
    "CYAN",
    "MAGENTA",
    "NEGRO",
    "BLANCO",
    "GRIS",
    "NARANJA"
]
clientes = {}
prestamo = {}
prestamo_por_retornar = {}
prestamo_por_periodo = {}
duracion_prestamos = {}
ranking_clientes = {}
lista_prestamo_reportes = []

df_prestamos = pd.DataFrame(lista_prestamo_reportes)
reporte_rodada = pd.DataFrame()
reporte_color = pd.DataFrame()

def mostrar_ruta():
    print(' > '.join(ruta_actual))

def validar_texto(texto):
    for x in texto:
        if x.isdigit():
            return False
    else:
        return True
    
def mostrar_unidades():
    print(f'''
{'=' * 21} UNIDADES {'=' * 20}  
{'=' * 51}
{'Clave':^5}  {'Rodada':^20}  {'Color':^10} {'Estado':^10}
{'=' * 51}
    ''')
    for clave, datos in unidades.items():
        print(f'{clave:^5}  {datos[0]:^20}  {datos[1]:^10}  {datos[2]:^10}')


def mostrar_clientes():
    print(f'''
{'=' * 19} REPORTE DE CLIENTES {'=' * 20}  
{'=' * 60}
{'Clave':^10} \t {'Apellidos':^10} \t {'Nombres':^10} \t {'Telefono':^10}
{'=' * 60}
    ''')
    for clave, datos in clientes.items():
        print(f'{clave:^10} \t {datos[0]:^10} \t {datos[1]:^10} \t {datos[2]:^10}')

    
def buscar_claves(clave, diccionario):
    if clave in diccionario:
        return True
    else:
        return False
    
def iniciar_bd():
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("CREATE TABLE IF NOT EXISTS unidades \
                            (idUnidad INTEGER PRIMARY KEY, \
                            rodada INTERGER NOT NULL, \
                            color TEXT NOT NULL, \
                            estadoprestamo TEXT NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS clientes \
                           (idCliente INTERGER PRIMARY KEY, \
                           apellidos TEXT NO NULL, \
                           nombres TEXT NOT NULL, \
                           telefono INTERGER NOT NULL);")
            cursor.execute("CREATE TABLE IF NOT EXISTS prestamos \
                           (idPrestamo INTERGER PRIMARY KEY, \
                           idUnidad INTERGER NOT NULL, \
                           idCliente INTERGER NOT NULL, \
                           fechaprestamo TEXT NOT NULL, \
                           diasprestamo INTERGER NOT NULL,\
                           fecharetorno TEXT, \
                           FOREIGN KEY (idUnidad) REFERENCES unidades(idUnidad), \
                           FOREIGN KEY (idCliente) REFERENCES clientes(idCliente));")
    except Error as e:
        print(e)
    finally:
        conexion.close()

def insertar_unidades_sqlite(valores):
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("INSERT INTO unidades VALUES(?,?,?,?)", valores)
    except Error as e:
        print(e)
    finally:
        conexion.close()

def insertar_clientes_sqlite(valores):
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("INSERT INTO clientes VALUES (?,?,?,?)", valores)
    except Error as e:
        print(e)
    finally:
        conexion.close()

def insertar_prestamos_sqlite(valores):
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("INSERT INTO prestamos VALUES (?,?,?,?,?,?)", valores)
    except Error as e:
        print(e)
    finally:
        conexion.close()

def actualizar_estado_prestamo(valores):
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("UPDATE unidades SET estadoprestamo = ? WHERE idUnidad = ?", valores)
    except Error as e:
        print(e)
    finally:
        conexion.close()

def actualizar_fecha_retorno(fecha_retorno):
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("UPDATE prestamos SET fecharetorno = ? WHERE idPrestamo = ?", fecha_retorno)
    except Error as e:
        print(e)
    finally:
        conexion.close()

def cargar_unidades():
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM unidades ORDER BY idUnidad")
            registros = cursor.fetchall()
            data = []
            if registros:
                for clave, rodada, color, prestamo in registros:
                    unidades[int(clave)] = [int(rodada), color, prestamo]
                    data.append([int(clave), int(rodada), color, prestamo])
                unidades_df = pd.DataFrame(data, columns = ["clave", "rodada", "color", "prestamo"])
                return unidades_df
            else:
                print('No hay unidades que cargar')
    except Error as e:
        print(e)
    finally:
        conexion.close()
        
def cargar_clientes():
    try:
        with sqlite3.connect("Tienda Bicicletas.db") as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM clientes ORDER BY idCliente")
            registros = cursor.fetchall()
            data = []
            if registros:
                for clave, apellidos, nombres, telefono in registros:
                    clientes[int(clave)] = (apellidos, nombres, telefono)
                    data.append([int(clave), apellidos, nombres, telefono])
                    clientes_df = pd.DataFrame(data, columns=['cliente_id', 'apellidos', 'nombres', 'telefono'])
                clientes_df['nombre'] = clientes_df['apellidos'] + ' ' + clientes_df['nombres']
                return clientes_df
            else:
                print('No hay clientes que cargar')
    except Error as e:
        print(e)
    finally:
        conexion.close()

def cargar_prestamos():
    try:
        with sqlite3.connect("Tienda Bicicletas.db", detect_types=sqlite3.PARSE_DECLTYPES | sqlite3.PARSE_COLNAMES) as conexion:
            cursor = conexion.cursor()
            cursor.execute("SELECT * FROM prestamos ORDER BY idPrestamo")
            registros = cursor.fetchall()
            data = []
            if registros:
                for folio, clave_unidad, clave_cliente, fecha_prestamo, dias_prestamo, fecha_retorno in registros:
                    fecha_prestamo = datetime.datetime.strptime(fecha_prestamo, '%Y-%m-%d').date()
                    if fecha_retorno != 'NULL':
                        fecha_retorno = datetime.datetime.strptime(fecha_retorno, '%Y-%m-%d').date()
                        prestamo[int(folio)] = [int(clave_unidad), int(clave_cliente), fecha_prestamo.strftime('%m-%d-%Y'), int(dias_prestamo), fecha_retorno.strftime('%m-%d-%Y')]
                    else:
                        prestamo[int(folio)] = [int(clave_unidad), int(clave_cliente), fecha_prestamo.strftime('%m-%d-%Y'), int(dias_prestamo), fecha_retorno]
                    data.append([int(folio), int(clave_unidad), int(clave_cliente), fecha_prestamo, int(dias_prestamo), fecha_retorno])
                
                prestamos_df = pd.DataFrame(data, columns=['prestamo_id', 'clave', 'cliente_id', 'fecha_prestamo', 'dias_prestamo', 'fecha_retorno'])
                
                prestamos_df['fecha_prestamo'] = pd.to_datetime(prestamos_df['fecha_prestamo'], format='%Y-%m-%d', errors='coerce')
                prestamos_df['fecha_retorno'] = pd.to_datetime(prestamos_df['fecha_retorno'], format='%Y-%m-%d', errors='coerce')

                prestamos_df = prestamos_df.dropna(subset=['fecha_prestamo', 'fecha_retorno'])  
                return prestamos_df
            else:
                print("No hay prestamos que cargar")
    except Error as e:
        print(e)
    finally:
        conexion.close()

def cargar_prestamo_por_retornar():
    for clave, datos in prestamo.items():
        if datos[4] == 'NULL':
            prestamo_por_retornar[int(clave)] = [int(datos[0]), int(datos[1]), datos[2], int(datos[3]), datos[4]]

def cargar_lista_prestamos():  
    for clave, datos in unidades.items():
        for folio, detalles in prestamo.items():
            if detalles[0] == clave:
                fecha_prestamo = datetime.datetime.strptime(detalles[2], '%m-%d-%Y').date()
                data = {
                    'Clave unidad': int(clave),
                    'Color': datos[1],
                    'Rodada': int(datos[0]),
                    'Fecha prestamo': fecha_prestamo  
                }
                lista_prestamo_reportes.append(data)

def exportar_csv_clientes(nombre_archivo, diccionario):
    with open(nombre_archivo, 'w', encoding='latin1', newline='') as archivo:
        grabador = csv.writer(archivo)
        grabador.writerow(('Clave', 'Apellidos', 'Nombres', 'Telefono'))
        grabador.writerows([(clave, datos[0], datos[1], datos[2]) for clave, datos in diccionario.items()])

def exportar_csv_cliente_especifico(nombre, lista):
    with open(nombre, 'w', encoding='latin1', newline='') as archivo:
        grabador = csv.writer(archivo)
        grabador.writerow(('Clave', 'Apellidos', 'Nombres', 'Telefono'))
        grabador.writerows([(datos[0], datos[1], datos[2], datos[3], datos[4], datos[5]) for datos in lista])

def exportar_csv_prestamo(nombre_archivo, diccionario):
    with open(nombre_archivo, 'w', encoding='latin1', newline='') as archivo:
        grabador = csv.writer(archivo)
        grabador.writerow(("Folio", "Clave Unidad", "Clave Cliente", "Fecha prestamo", "Dias de prestamo", "Fecha de retorno"))
        grabador.writerows([(folio, datos[0], datos[1], datos[2], datos[3], datos[4]) for folio, datos in diccionario.items()])

def exportar_csv_reportes_prestamos(nombre_archivo, diccionario):
    with open(nombre_archivo, 'w', encoding='latin1', newline='') as archivo:
        grabador = csv.writer(archivo)
        grabador.writerow(("Clave Unidad", "Rodada", "Fecha prestamo", "Nombres", "Apellidos", "Telefono"))
        grabador.writerows([(datos[0], unidades[datos[0]][0], datos[2], clientes[datos[1]][1], clientes[datos[1]][0], clientes[datos[1]][2]) for folio, datos in diccionario.items()])

def exportar_clientes_excel():
    libro = openpyxl.Workbook()
    hoja = libro["Sheet"]
    hoja.title = "Clientes"

    hoja["A1"].value = "Clave"
    hoja["B1"].value = "Apellidos"
    hoja["C1"].value = "Nombres"
    hoja["D1"].value = "Telefono"

    fila = 2
    for clave, datos in clientes.items():
        hoja.cell(row=fila, column=1).value = clave
        hoja.cell(row=fila, column=2).value = datos[0]
        hoja.cell(row=fila, column=3).value = datos[1]
        hoja.cell(row=fila, column=4).value = datos[2]
        fila += 1

    libro.save("Clientes.xlsx")
    print('\nArchivos exportados a excel exitosamente!!')

def exportar_cliente_especifico_excel(nombre, lista):
    libro = openpyxl.Workbook()
    hoja2 = libro["Sheet"]
    hoja2.title = "Por Retornar"

    hoja2["A1"].value = "Folio"
    hoja2["B1"].value = "Clave Unidad"
    hoja2["C1"].value = "Clave Cliente"
    hoja2["D1"].value = "Fecha de prestamo"
    hoja2["E1"].value = "Dias de prestamo"
    hoja2["F1"].value = "Fecha de retorno"

    fila = 2
    for datos in lista:
        hoja2.cell(row=fila, column=1).value = datos[0]
        hoja2.cell(row=fila, column=2).value = datos[1]
        hoja2.cell(row=fila, column=3).value = datos[2]
        hoja2.cell(row=fila, column=4).value = datos[3]
        hoja2.cell(row=fila, column=5).value = datos[4]
        hoja2.cell(row=fila, column=6).value = datos[5]
        fila += 1

    libro.save(f"{nombre}.xlsx")
    print('\nArchivos exportados a excel exitosamente!!')

def exportar_rpr_excel():
    libro = openpyxl.Workbook()
    hoja2 = libro["Sheet"]
    hoja2.title = "Por Retornar"

    hoja2["A1"].value = "Folio"
    hoja2["B1"].value = "Clave Unidad"
    hoja2["C1"].value = "Clave Cliente"
    hoja2["D1"].value = "Fecha de prestamo"
    hoja2["E1"].value = "Dias de prestamo"
    hoja2["F1"].value = "Fecha de retorno"

    fila = 2
    for clave, datos in prestamo_por_retornar.items():
        hoja2.cell(row=fila, column=1).value = clave
        hoja2.cell(row=fila, column=2).value = unidades[datos[0]][0]
        hoja2.cell(row=fila, column=3).value = datos[2]
        hoja2.cell(row=fila, column=4).value = clientes[datos[1]][1]
        hoja2.cell(row=fila, column=5).value = clientes[datos[1]][0]
        hoja2.cell(row=fila, column=6).value = clientes[datos[1]][2]
        fila += 1

    libro.save("Por Retornar.xlsx")
    print('\nArchivos exportados a excel exitosamente!!')

def exportar_rpp_excel():
    libro = openpyxl.Workbook()
    hoja2 = libro["Sheet"]
    hoja2.title = "Por Periodo"

    hoja2["A1"].value = "Folio"
    hoja2["B1"].value = "Clave Unidad"
    hoja2["C1"].value = "Clave Cliente"
    hoja2["D1"].value = "Fecha de prestamo"
    hoja2["E1"].value = "Dias de prestamo"
    hoja2["F1"].value = "Fecha de retorno"

    fila = 2
    for clave, datos in prestamo_por_periodo.items():
        hoja2.cell(row=fila, column=1).value = clave
        hoja2.cell(row=fila, column=2).value = unidades[datos[0]][0]
        hoja2.cell(row=fila, column=3).value = datos[2]
        hoja2.cell(row=fila, column=4).value = clientes[datos[1]][1]
        hoja2.cell(row=fila, column=5).value = clientes[datos[1]][0]
        hoja2.cell(row=fila, column=6).value = clientes[datos[1]][2]
        fila += 1

    libro.save("Por Periodo.xlsx")
    print('\nArchivos exportados a excel exitosamente!!')

def registrar_unidad():
    while True:
        clave_unidad = max(unidades, default=0) + 1
        rodada_unidad = input('Que tipo de rodada desea registrar (20, 26, 29): ')
        if rodada_unidad.isdigit():
            rodada_unidad = int(rodada_unidad)
            if rodada_unidad == 20:
                break
            elif rodada_unidad == 26:
                break
            elif rodada_unidad == 29:
                break
            else:
                print('Ese tipo de rodada no esta disponible')
                continue
        else:
            print('La rodad debe ser un  caracter de tipo numerico')
            continue
    while True:
        print(f'''\n
{'=' * 12} LISTA DE COLORES {'=' * 12} 
{'=' * 42}
''')
        for color in colores:
            print(f'{color:^42}')
            
        color_unidad = input('\nDe la lista de colores, ingresa el color deseado: ')
        if not color_unidad:
            print('\nLa unidad debe ser de un color obligatoriamente.')
            continue
        elif len(color_unidad) > 15:
            print('\nEl color no puede superar los 15 caracteres.')
            continue
        elif color_unidad.isdigit():
            print('\nEl codigo de color debe ser un caracter de texto.')
            continue
        elif not color_unidad.upper() in colores:
            print('No exite un color registrado con ese nombre.')
            continue
        else:
            prestada = 'DISPONIBLE'
            unidades[clave_unidad] = [rodada_unidad, color_unidad.upper(), prestada]
            valores_unidades = (clave_unidad, rodada_unidad, color_unidad.upper(), prestada)
            insertar_unidades_sqlite(valores_unidades)
            break
             
    print('\nUnidad registrada con exito!')

def registrar_clientes():
    clave_cliente = max(clientes, default=0) +1
    while True:
        apellidos = input('Ingrese los apellidos del cliente: ')
        if not apellidos:
            print('No se permiten campos vacios.\n')
            continue
        elif len(apellidos)>40:
            print('No se permite ingresar mas de 40 caracteres.\n')
            continue
        elif validar_texto(apellidos):
            break
        else:
            print('No se permiten caracteres numericos.')
            continue
    while True:
        nombres = input('Ingresa los nombres del cliente: ')
        if not nombres:
            print('No se permiten campos vacios.\n')
            continue
        elif len(nombres) > 40:
            print('No se permite ingresar mas de 40 caracteres.\n')
            continue
        elif validar_texto(nombres):
            break
        else:
            print('No se permiten caracteres numericos.')
            continue
    while True:
        telefono = input('Ingresa el telefono del cliente(10 digitos): ')
        if not telefono:
            print('No se permiten campos vacios.\n')
            continue
        elif validar_texto(telefono) == False and len(telefono) == 10:
            break
        else:
            print('El numero solo pude tener carateres numericos y 10 digitos.')
            continue
    clientes[clave_cliente] = (apellidos, nombres, telefono)
    valores_cliente = (clave_cliente, apellidos, nombres, telefono)
    insertar_clientes_sqlite(valores_cliente)
    print('\nCliente registrado exitosamente!')

def registrar_prestamo():
    folio_prestamo = max(prestamo, default=0) + 1
    salir = False
    while True:
        mostrar_unidades()
        clave_unidad = input('\nIngresa la clave de unidad para el prestamo.\nSi no hay unidades disponibles deja el campo vacio:  ')
        if not clave_unidad:
            salir = True 
            break
        elif not clave_unidad.isdigit():
            print('\nLas claves solo pueden ser caracteres numéricos enteros.')
            continue
        
        clave_unidad = int(clave_unidad)
        
        if clave_unidad not in unidades:
            print(f'\nLa clave {clave_unidad} no está registrada en el sistema.')
            continue
        elif unidades[clave_unidad][2] == 'PRESTADA':
            print('\nUnidad se encuentra en préstamo. Intente con otra.')
            continue
        
        print('\nUnidad encontrada')
        break

    while True:
        if salir:
            break
        mostrar_clientes()
        clave_cliente = input('\nIngresa la clave del cliente para el prestamo: ')
        if clave_cliente.isdigit():
            clave_cliente = int(clave_cliente)
            if clave_cliente < 1:
                print('\nNo hay unidades con claves menores a 0 asignadas.')
                continue
            elif buscar_claves(clave_cliente, clientes):
                print('\nCliente encontrado')
                break
            elif buscar_claves(clave_cliente, clientes) == False:
                print(f'\nLa clave {clave_cliente} no esta registrada en el sistema.')
                continue
        elif not clave_cliente:
                print('\nNo se permiten campos vacios.')
                continue
        else:
            print('\nLas claves solo pueden ser caracteres numericos enteros.')
            continue
    while True:
        if salir:
            break
        fecha_prestamo = input('\nIngresa la fecha de prestamo (mm-dd-aaaa) dejar vacio para establecer la fecha actual: ')
        fecha_actual = datetime.date.today()
        if not fecha_prestamo:
            fecha_prestamo = fecha_actual
            print('\nSe registro el prestamo con la fecha actual')
        else:
            try:
                fecha_prestamo = datetime.datetime.strptime(fecha_prestamo, "%m-%d-%Y").date()
                print(f'\nSe registro el prestamo con la fecha {fecha_prestamo.strftime("%m-%d-%Y")}')
            except ValueError:
                print("\nFormato de fecha incorrecto. Inténtalo de nuevo.")
                continue
        
        if fecha_prestamo < fecha_actual:
            print('\nNo se permiten fechas menores a la actual.')
            continue
        else:
            break
    while True:
        if salir:
            break
        cant_dias = input('\nIngrese la cantidad de dias de prestamo (1-14): ')
        if not cant_dias:
            print('\nNo se permiten campos vacios.')
            continue
        elif cant_dias.isdigit():
            cant_dias = int(cant_dias)
            if cant_dias >= 1 and cant_dias <= 14:
                print(f'\nLa cantidad de dias de prestamo de la unidad {clave_unidad} es {cant_dias}')
                fecha_retorno = "NULL"
                unidades[clave_unidad][2] = 'PRESTADA'
                valores_unidad = ("PRESTADA", clave_unidad)
                actualizar_estado_prestamo(valores_unidad)
                prestamo[folio_prestamo] = [clave_unidad, clave_cliente, fecha_prestamo.strftime("%m-%d-%Y"), cant_dias, fecha_retorno]
                prestamo_por_retornar[folio_prestamo] = [clave_unidad, clave_cliente, fecha_prestamo.strftime("%m-%d-%Y"), cant_dias, fecha_retorno]
                lista_prestamo_reportes.append({'Clave unidad': clave_unidad, 'Color': unidades[clave_unidad][1], 'Rodada': unidades[clave_unidad][0], 'Fecha prestamo': fecha_prestamo.strftime("%m-%d-%Y")})
                valores_prestamo = (folio_prestamo, clave_unidad, clave_cliente, fecha_prestamo, cant_dias, fecha_retorno)
                insertar_prestamos_sqlite(valores_prestamo)
                print('\nPrestamo registrado con exito!')
                break
            else:
                print('\nMinimo se debe de prestar la unidad 1 dia y maximo 14')
                continue
        else:
            print('\nSolo se permiten valores numericos.')
            continue

def calcular_duracion_prestamos():
    if prestamo:
        df_prestamos = pd.DataFrame.from_dict(prestamo, orient='index', 
                                           columns=['Clave Unidad', 'Clave Cliente', 
                                                    'Fecha Prestamo', 'Dias Prestamo', 
                                                    'Fecha Retorno'])

        media = df_prestamos['Dias Prestamo'].mean()
        mediana = df_prestamos['Dias Prestamo'].median()
        moda = df_prestamos['Dias Prestamo'].mode()
        moda = moda.iloc[0] if not moda.empty else None
        minimo = df_prestamos['Dias Prestamo'].min()
        maximo = df_prestamos['Dias Prestamo'].max()
        desviacion_estandar = df_prestamos['Dias Prestamo'].std()
        q1 = df_prestamos['Dias Prestamo'].quantile(0.25)

        duracion_prestamos = {
            "Media": media,
            "Mediana": mediana,
            "Moda": moda,
            "Mínimo": minimo,
            "Máximo": maximo,
            "Desviación Estándar": desviacion_estandar,
            "Cuartil 1 (Q1)": q1
        }
        
        print("\n Estadísticas de Duración de Préstamos:")
        print(f"Media: {media}")
        print(f"Mediana: {mediana}")
        print(f"Moda: {moda}")
        print(f"Minimo: {minimo}")
        print(f"Maximo: {maximo}")
        print(f"Desviación Estandar: {desviacion_estandar}")
        print(f"Cuartil 1 (Q1): {q1}")
    else:
        print("\nNo hay préstamos para calcular estadísticas.")

def retornar():
    if prestamo and prestamo_por_retornar:
        salir = False
        while salir == False:
            prestamos_ordenados = sorted(prestamo_por_retornar.items(), key=lambda x: datetime.datetime.strptime(x[1][2], "%m-%d-%Y"), reverse=True)
            print(f'''
{'=' * 36} REPORTE POR RETORNAR {'=' * 36}  
{'=' * 94}
{'Folio':^10}  {'Clave Unidad':^10}  {'Clave Cliente':^10}  {'Fecha de prestamo':^10}  {'Dias de prestamo':^10}  {'Fecha de retorno':^10}
{'=' * 94}
    ''')
            prestamos_ordenados = sorted(prestamo_por_retornar.items(), key=lambda x: datetime.datetime.strptime(x[1][2], '%m-%d-%Y'), reverse=True)
            for clave, datos in prestamos_ordenados:
                print(f'{clave:^10}   {datos[0]:^10}     {datos[1]:^10}       {datos[2]:^10}         {datos[3]:^10}       {datos[4]:^10}')
                
            folio_a_retornar = input('\nIngresa el folio del préstamo que deseas retornar: ')
            if not folio_a_retornar:
                print('\nNo se permiten campos vacíos.')
                continue
            elif folio_a_retornar.isdigit():
                folio_a_retornar = int(folio_a_retornar)
                folio_encontrado = False  
                for clave, datos in prestamo.items():
                    if clave == folio_a_retornar:
                        folio_encontrado = True
                        print('\n¡Folio de préstamo encontrado!')
                        print(f'Folio: {clave}')
                        print(f'Clave unidad: {datos[0]}')
                        print(f'Clave cliente: {datos[1]}')
                        print(f'Fecha préstamo: {datos[2]}')
                        print(f'Dias de préstamo: {datos[3]}')
                        print(f'Fecha retorno: {datos[4]}')
                        
                        while True:
                            nueva_fecha_retorno = input(f'\nIntroduce la fecha de retorno de la unidad {datos[0]} (mm-dd-aaaa): ')
                            fecha_prestamo = datetime.datetime.strptime(datos[2], "%m-%d-%Y").date()

                            if not nueva_fecha_retorno:
                                print('\nFecha de retorno no establecida. Regresando al menú.')
                                salir = True
                                break 
                            else:
                                try:
                                    nueva_fecha_retorno = datetime.datetime.strptime(nueva_fecha_retorno, "%m-%d-%Y").date()
                                except ValueError:
                                    print('\nFormato incorrecto. Vuelve a intentarlo.')
                                    continue

                            if nueva_fecha_retorno < fecha_prestamo:
                                print('\nLa fecha de retorno no puede ser anterior a la fecha de préstamo.')
                                continue
                            elif nueva_fecha_retorno > (fecha_prestamo + datetime.timedelta(days=datos[3])):
                                datos[4] = nueva_fecha_retorno.strftime("%m-%d-%Y")
                                prestamo_por_retornar[clave][4] = nueva_fecha_retorno.strftime("%m-%d-%Y")
                                print('\nEl retorno superó los días de préstamo.')
                                unidades[datos[0]][2] = 'DISPONIBLE'
                                valores_unidad = ("DISPONIBLE", datos[0])
                                actualizar_estado_prestamo(valores_unidad)
                                print('\nFecha de retorno actualizada con éxito.')
                                del prestamo_por_retornar[folio_a_retornar]
                                valores = (nueva_fecha_retorno, folio_a_retornar)
                                actualizar_fecha_retorno(valores)
                                salir = True 
                                break 
                            else:
                                datos[4] = nueva_fecha_retorno.strftime("%m-%d-%Y")
                                prestamo_por_retornar[clave][4] = nueva_fecha_retorno.strftime("%m-%d-%Y")
                                unidades[datos[0]][2] = 'DISPONIBLE'
                                valores_unidad = ("DISPONIBLE", datos[0])
                                actualizar_estado_prestamo(valores_unidad)
                                print('\nFecha de retorno actualizada con éxito.')
                                salir = True
                                del prestamo_por_retornar[folio_a_retornar]
                                valores = (nueva_fecha_retorno, folio_a_retornar)
                                actualizar_fecha_retorno(valores)
                                break  
                        break  
                if not folio_encontrado:
                    print(f'\nNo se encontró ningún folio con la clave {folio_a_retornar}')
                    continue 
            else:
                print('\nEl folio debe ser un carácter numérico entero.')
                continue
    else:
        print('\nNo hay préstamos pendientes por retornar.')

def reporte_ranking_clientes():
    try:
        df_prestamos = pd.DataFrame.from_dict(prestamo, orient='index', columns=['Clave Unidad', 'Clave Cliente', 'Fecha Prestamo', 'Dias Prestamo', 'Fecha Retorno'])
        df_clientes = pd.DataFrame.from_dict(clientes, orient='index', columns=['Apellidos', 'Nombres', 'Telefono']) 

        df_prestamos = df_prestamos.rename(columns={'Clave Cliente': 'clave'})

        rentas_por_cliente = df_prestamos.groupby('clave').size().reset_index(name='Rentas')

        df_ranking = pd.merge(rentas_por_cliente, df_clientes, left_on='clave', right_index=True)

        df_ranking['Nombre Completo'] = df_ranking['Apellidos'] + ' ' + df_ranking['Nombres']

        df_ranking = df_ranking.sort_values('Rentas', ascending=False)

        print(f'''
        {'=' * 65} 'Ranking de Clientes:' {'=' * 60}
        {'=' * 150}
        ''')
        for index, row in df_ranking.iterrows():
            print(f"Rentas: {row['Rentas']}, Clave: {row['clave']}, Nombre: {row['Nombre Completo']}, Teléfono: {row['Telefono']}")

    except KeyError as e:
        print(f"\nError: La columna {e} no existe. Verifica los nombres de las columnas en los DataFrames.")
    except Exception as e:
        print(f"\nOcurrió un error inesperado: {e}")

def reporte_clientes():
    if clientes: 
        print(f'''
{'=' * 19} REPORTE DE CLIENTES {'=' * 20}  
{'=' * 60}
{'Clave':^10} \t {'Apellidos':^10} \t {'Nombres':^10} \t {'Telefono':^10}
{'=' * 60}
    ''')
        for clave, datos in clientes.items():
            print(f'{clave:^10} \t {datos[0]:^10} \t {datos[1]:^10} \t {datos[2]:^10}')
        
        while True:
            opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
            if not opcion_exportar:
                break
            elif opcion_exportar.upper() == 'CSV':
                exportar_csv_clientes('Reporte Clientes.csv', clientes)
                break
            elif opcion_exportar.upper() == 'MSEXCEL':
                exportar_clientes_excel()
                break
            else:
                print ('\nOpcion no valida')
                continue
    else:
        print('\nNo hay clientes para mostrar')

def reporte_cliente_especifico():
    if clientes:
        while True:
            cliente_consultar = input('Ingresa la clave del cliente que desea consultar: ')
            if cliente_consultar.isdigit():
                cliente_consultar = int(cliente_consultar)
                for clave, datos in clientes.items():
                    if clave == cliente_consultar:
                        print(f'\nClave de cliente: {clave}')
                        print(f'Apellido(s): {datos[0]}')
                        print(f'Nombre(s): {datos[1]}')
                        print(f'Telefono: {datos[2]}')
                        
                        prestamos_cliente = []
                        devoluciones_cliente = []
                        
                        for folio, data in prestamo.items():
                            if data[1] == clave and data[4] == 'NULL':
                                prestamos_cliente.append([folio, data[0], data[1], data[2], data[3], data[4]])
                        
                        if prestamos_cliente:
                            print(f'''
{'=' * 84 }
{'=' * 25} Reporte De Prestamos De {datos[1]} {'=' * 25}
{'='*84}
{'Folio:':^10} {'Unidad':^10} {'Cliente':^10} {'Fecha de prestamo':^17} {'Dias de prestamo':^16} {'Fecha de Retorno':^17}
{'=' * 84}
''')
                            for prestamo_data in prestamos_cliente:
                                print(f'{prestamo_data[0]:^10} {prestamo_data[1]:^10} {prestamo_data[2]:^10} {prestamo_data[3]:^17} {prestamo_data[4]:^16} {prestamo_data[5]:^17}')
                        
                            while True:
                                opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
                                if not opcion_exportar:
                                    break
                                elif opcion_exportar.upper() == 'CSV':
                                    exportar_csv_cliente_especifico(f'Prestamos de {datos[1]}.csv', prestamos_cliente)
                                    break
                                elif opcion_exportar.upper() == 'MSEXCEL':
                                    exportar_cliente_especifico_excel(f'Prestamos de {datos[1]}.xlsx', prestamos_cliente)
                                    break
                                else:
                                    print ('\nOpcion no valida')
                                    continue

                        for folio, data in prestamo.items():
                            if data[1] == clave and data[4] != 'NULL':
                                devoluciones_cliente.append([folio, data[0], data[1], data[2], data[3], data[4]])
                        
                        if devoluciones_cliente:
                            print(f'''\n
{'=' * 84 }
{'=' * 25} Reporte De Devoluciones De {datos[1]} {'=' * 25}
{'='*84}
{'Folio:':^10} {'Unidad':^10} {'Cliente':^10} {'Fecha de prestamo':^17} {'Dias de prestamo':^16} {'Fecha de Retorno':^17}
{'=' * 84}
''')
                            for devolucion_data in devoluciones_cliente:
                                print(f'{devolucion_data[0]:^10} {devolucion_data[1]:^10} {devolucion_data[2]:^10} {devolucion_data[3]:^17} {devolucion_data[4]:^16} {devolucion_data[5]:^17}')
                        
                            while True:
                                opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
                                if not opcion_exportar:
                                    break
                                elif opcion_exportar.upper() == 'CSV':
                                    exportar_csv_cliente_especifico(f'Devoluciones de {datos[1]}.csv', devoluciones_cliente)
                                    break
                                elif opcion_exportar.upper() == 'MSEXCEL':
                                    exportar_cliente_especifico_excel(f'Devoluciones de {datos[1]}.xlsx', devoluciones_cliente)
                                    break
                                else:
                                    print ('\nOpcion no valida')
                                    continue

                        if not prestamos_cliente and not devoluciones_cliente:
                            print(f'\nNo hay préstamos o devoluciones asociados a la clave del cliente {cliente_consultar}.\n')
                        
                        break
            elif not cliente_consultar:
                salir = input('No puedes dejar campos vacios. Deseas regresar el menu de reportes?(SI/NO): ')
                if salir.upper() == 'SI':
                    break
                elif salir.upper() == 'NO':
                    continue 
            else:
                print('Solo se permiten claves numericas.')
            break
    else:
        print('\nNo hay clientes para consultar.')


def reporte_por_retornar():
    if prestamo_por_retornar:
        while True:
            fecha1 = input('\nIngresa la fecha de inicio del reporte (mm-dd-aaaa): ')
            fecha2 = input('\nIngresa la fecha final del reporte (mm-dd-aaaa): ')
            if not fecha1 or not fecha2: 
                print('\nLas fechas no pueden ser campos vacios.')
                continue
            else:
                try:
                    fecha1 = datetime.datetime.strptime(fecha1, "%m-%d-%Y").date()
                    fecha2 = datetime.datetime.strptime(fecha2, "%m-%d-%Y").date()
                except ValueError:
                    print('Formato incorrecto. Volver a intentar.')
                    continue

            if fecha1 > fecha2:
                print('\nNo se permite que la fecha final sea menor a la fecha inicial.')
            else:
                for clave, datos in prestamo.items():
                    if datos[4] == '--':
                        prestamo_por_retornar[clave] = datos
                print(f'''
{'=' * 27} REPORTE POR RETORNAR {'=' * 27}  
{'=' * 75}
{'Clave Unidad':^10}  {'Rodada':^6}  {'Fecha de prestamo':^10}  {'Nombre Completo':^20}  {'Telefono':^10}
{'=' * 75}
    ''')
                for clave, datos in prestamo_por_retornar.items():
                    print(f'{datos[0]:^10}   {unidades[datos[0]][0]:^6}       {datos[2]:^10}         {clientes[datos[1]][1]} {clientes[datos[1]][0]}       {clientes[datos[1]][2]:^10}')
                
                while True:
                    opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
                    if not opcion_exportar:
                        break
                    elif opcion_exportar.upper() == 'CSV':
                        exportar_csv_prestamo('Reporte Por Retornar.csv', prestamo_por_retornar)
                        exportar_csv_reportes_prestamos('Reporte Por Retornar 1.csv', prestamo_por_retornar)
                        break
                    elif opcion_exportar.upper() == 'MSEXCEL':
                        exportar_rpr_excel()
                        break
                    else:
                        print ('\nOpcion no valida')
                        continue
                break
    else:
        print('\nNo hay prestamos que retornar.')

def reporte_por_periodo():
    if prestamo:
        while True:
            fecha1 = input('\nIngresa la fecha de inicio del reporte (mm-dd-aaaa): ')
            fecha2 = input('\nIngresa la fecha final del reporte (mm-dd-aaaa): ')
            if not fecha1 or not fecha2: 
                print('\nLas fechas no pueden ser campos vacios.')
                continue
            else:
                try:
                    fecha1 = datetime.datetime.strptime(fecha1, "%m-%d-%Y").date()
                    fecha2 = datetime.datetime.strptime(fecha2, "%m-%d-%Y").date()
                except ValueError:
                    print('Formato incorrecto. Volver a intentar.')
                    continue

            if fecha1 > fecha2:
                print('\nNo se permite que la fecha final sea menor a la fecha inicial.')
                continue
            else:
                print(f'''
{'=' * 27} REPORTE POR PERIODO {'=' * 28}  
{'=' * 75}
{'=' * 21} PERIODO {fecha1.strftime("%m-%d-%Y")} a {fecha2.strftime("%m-%d-%Y")} {'=' * 21}
{'Clave Unidad':^10}  {'Rodada':^6}  {'Fecha de prestamo':^10}  {'Nombre Completo':^10}  {'Telefono':^10}
{'=' * 75}
    ''')
                for clave, datos in prestamo.items():
                    fecha_prestamo = datetime.datetime.strptime(datos[2], "%m-%d-%Y").date()
                    if fecha_prestamo >= fecha1 and fecha_prestamo <= fecha2:
                        print(f'{datos[0]:^10}   {unidades[datos[0]][0]:^6}       {datos[2]:^10}         {clientes[datos[1]][1]} {clientes[datos[1]][0]}       {clientes[datos[1]][2]:^10}')
                        prestamo_por_periodo[clave] = datos

                while True:
                    opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
                    if not opcion_exportar:
                        break
                    elif opcion_exportar.upper() == 'CSV':
                        exportar_csv_prestamo('Reporte Por Periodo.csv', prestamo_por_periodo)
                        exportar_csv_reportes_prestamos('Reporte Por Periodo 1.csv', prestamo_por_periodo)
                        break
                    elif opcion_exportar.upper() == 'MSEXCEL':
                        exportar_rpp_excel()
                        break
                    else:
                        print ('\nOpcion no valida')
                        continue
                
                break
    else:
        print('\nNo hay prestamos que reportar')

def listado_por_rodada(unidades_df, rodada):
    return unidades_df[unidades_df['rodada'] == rodada][['clave', 'color']]

def listado_por_color(unidades_df, color):
    return unidades_df[unidades_df['color'] == color][['clave', 'rodada']]

def retrasos():
    print(f'''
{'=' * 37} REPORTE DE RETRASOS  {'=' * 36}  
{'=' * 94}
{'Dias de retraso':^10}  {'Fecha limite':^6}  {'Clave Unidad':^5}  {'Rodada':^6}  {'Color':^10}  {'Nombre Completo':^20} {'Telefono':^10}
{'=' * 94}
''')
    data = []
    for clave, datos in prestamo.items():
        hoy = datetime.date.today()
        fecha_prestamo = datetime.datetime.strptime(datos[2], "%m-%d-%Y").date()
        fecha_limite = (fecha_prestamo + datetime.timedelta(days=datos[3]))
        if datos[4] == 'NULL':
            dias_retraso = (hoy  - fecha_limite).days
        else:
            fecha_retorno = datetime.datetime.strptime(datos[4], "%m-%d-%Y").date()
            dias_retraso = (fecha_retorno - fecha_limite).days
        if dias_retraso > 0:
            print(f'{dias_retraso:^15}    {fecha_limite}      {datos[0]:^5}     {unidades[datos[0]][0]:^6} {unidades[datos[0]][1]:^10}       {clientes[datos[1]][1]} {clientes[datos[1]][0]}     {clientes[datos[1]][2]:^10}')
            data.append([dias_retraso, fecha_limite, datos[0], unidades[datos[0]][0], unidades[datos[0]][1], clientes[datos[1]][1], clientes[datos[1]][0], clientes[datos[1]][2]])
    while True:
        opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
        if not opcion_exportar:
            break
        elif opcion_exportar.upper() == 'CSV':
            with open('Reporte retrasos.csv', 'w', encoding='latin1', newline='') as archivo:
                grabador = csv.writer(archivo)
                grabador.writerow(("Dias de retraso", "Fecha limite", "Clave Unidad", "Rodada", "Color", "Nombre", "Apellido", "Telefono"))
                for datos in data:
                    grabador.writerow((datos))
            break
        elif opcion_exportar.upper() == 'MSEXCEL':
            libro = openpyxl.Workbook()
            hoja2 = libro["Sheet"]
            hoja2.title = "Retrasos"

            hoja2["A1"].value = "Dias de retraso"
            hoja2["B1"].value = "Fecha limite"
            hoja2["C1"].value = "Clave Unidad"
            hoja2["D1"].value = "Rodada"
            hoja2["E1"].value = "Color"
            hoja2["F1"].value = "Nombre"
            hoja2["G1"].value = "Apellido"
            hoja2["H1"].value = "Telefono"

            fila = 2
            for datos in data:
                hoja2.cell(row=fila, column=1).value = datos[0]
                hoja2.cell(row=fila, column=2).value = datos[1]
                hoja2.cell(row=fila, column=3).value = datos[2]
                hoja2.cell(row=fila, column=4).value = datos[3]
                hoja2.cell(row=fila, column=5).value = datos[4]
                hoja2.cell(row=fila, column=6).value = datos[5]
                hoja2.cell(row=fila, column=7).value = datos[6]
                hoja2.cell(row=fila, column=8).value = datos[7]
                fila += 1

            libro.save("Retrasos.xlsx")
            print('\nArchivos exportados a excel exitosamente!!')
            break
        else:
            print ('\nOpcion no valida')
            continue


def reporte_preferencia_rodada():
    df_prestamos = pd.DataFrame.from_dict(unidades, orient='index', columns=['Rodada', 'Color', 'Prestada'])
    reporte_rodada = df_prestamos.groupby('Rodada').size().reset_index(name='cantidad_prestamos')
    reporte_rodada = reporte_rodada.sort_values(by='cantidad_prestamos', ascending=False)
    
    print(reporte_rodada.to_string(index=False))
    return reporte_rodada

def graficar_prestamos_rodada(reporte_rodada):
  plt.pie(reporte_rodada['cantidad_prestamos'], labels=reporte_rodada['Rodada'], autopct='%1.1f%%', startangle=90)
  plt.axis('equal')   
  plt.title('Proporción de Préstamos por Rodada')
  plt.show()

def reporte_preferencia_color():
    df_prestamos = pd.DataFrame.from_dict(unidades, orient='index', columns=['Rodada', 'Color', 'Prestada'])
    reporte_color = df_prestamos.groupby('Color').size().reset_index(name='cantidad_prestamos')
    reporte_color = reporte_color.sort_values(by='cantidad_prestamos', ascending=False)
    
    print(reporte_color.to_string(index=False))
    return reporte_color

def graficar_prestamos_color(reporte_color):
  plt.pie(reporte_color['cantidad_prestamos'], labels=reporte_color['Color'], autopct='%1.1f%%', startangle=90)
  plt.axis('equal')  
  plt.title('Proporción de Préstamos por Color')
  plt.show()

def prestamos_por_dia():
  df_prestamos = pd.DataFrame.from_dict(prestamo, orient='index', columns=['Clave Unidad', 'Clave Cliente', 'Fecha Prestamo', 'Dias Prestamo', 'Fecha Retorno'])
  df_prestamos['Dia Semana'] = df_prestamos['Fecha Prestamo'].apply(lambda x: datetime.datetime.strptime(x, "%m-%d-%Y").date().weekday())
  prestamos_dia = df_prestamos.groupby('Dia Semana').size().reset_index(name='Cantidad Prestamos')

  prestamos_dia['Dia Semana'] = (prestamos_dia['Dia Semana'] + 1) % 7
  prestamos_dia = prestamos_dia.sort_values(by='Dia Semana')
  print(prestamos_dia.to_string(index=False))
  return prestamos_dia
    
def graficar_prestamos_dia(prestamos_dia):
    df_prestamos = pd.DataFrame(prestamos_dia)
    prestamos_por_dia = df_prestamos.groupby('Dia Semana').size().reset_index(name='cantidad_prestamos')
    dias_completos = pd.DataFrame({'Dia Semana': range(7)})  
    prestamos_por_dia = dias_completos.merge(prestamos_por_dia, on='Dia Semana', how='left').fillna(0)
    prestamos_por_dia = prestamos_por_dia.sort_values(by='Dia Semana')
    dias_semana = ['Domingo', 'Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
    
    plt.bar(prestamos_por_dia['cantidad_prestamos'], prestamos_por_dia['Dia Semana'], tick_label=dias_semana)
    plt.xlabel('Día de la Semana')
    plt.ylabel('Cantidad de Préstamos')
    plt.title('Préstamos Totales por Día de la Semana')
    plt.show()



iniciar_bd()
cargar_unidades()
cargar_clientes()
cargar_prestamos()
cargar_prestamo_por_retornar()
cargar_lista_prestamos()




while True:
    print('')
    mostrar_ruta()
    print(f'''
{'=' * 15}MENU PRINCIPAL{'=' * 15}
1.- REGISTRO
2.- REGISTRAR PRESTAMO
3.- RETORNO
4.- INFORMES
5.- SALIR
''')
    menu_principal = input('Ingresa la opcion deseada: ')
    if menu_principal.isdigit():
        menu_principal = int(menu_principal)
        if menu_principal == 1:
            ruta_actual.append('Menú Registros')
            while True:
                print('')
                mostrar_ruta()
                print(f'''\n
{'=' * 15}MENU DE REGISTRO{'=' * 15}
1.-REGISTRA UNIDAD
2.-REGISTRAR CLIENTE
3.-SALIR 
    ''')
                menu_registro = input('Ingresa la opcion deseada: ')
                if menu_registro.isdigit():
                    menu_registro = int(menu_registro)
                    if menu_registro == 1:
                        registrar_unidad()
                    elif menu_registro == 2:
                        registrar_clientes()
                    elif menu_registro == 3:
                        break
                    else:
                        print('Opcion no valida.')
                        continue
                else:
                    print('Solo se permiten opciones numericas.')
                    continue
            ruta_actual.pop()
        elif menu_principal == 2:
            if not unidades:
                print('No hay unidades registradas para pestamo')
                continue
            elif not clientes:
                print('No hay clientes registrados para prestamo.')
                continue
            else:
                registrar_prestamo()
        elif menu_principal == 3:
            retornar()
        elif menu_principal == 4:
            ruta_actual.append('Informes')
            while True:
                print('')
                mostrar_ruta()
                print(f'''\n
{'=' * 15}MENU DE REPORTES{'=' * 15}
1.-REPORTES
2.-ANALISIS
3.-VOLVER AL MENU PRINCIPAL
    ''')
                menu_informes = input('Ingresa la opcion deseada: ')
                if menu_informes.isdigit():
                    menu_informes = int(menu_informes)
                    if menu_informes == 1:
                        ruta_actual.append('Reportes')
                        while True:
                            print('')
                            mostrar_ruta()
                            print(f'''\n
{'=' * 15}MENU DE REPORTES{'=' * 15}
1.-CLIENTES
2.-CLIENTE ESPECIFICO
3.-LISTADO DE UNIDADES
4.-RETRASOS
5.-PRESTAMOS POR RETORNAR
6.-PRESTAMOS POR PERIODO
7.-VOLVER AL MENU DE INFORMES
    ''')
                            menu_reportes = input('Ingresa la opcion deseada: ')
                            if menu_reportes.isdigit():
                                menu_reportes = int(menu_reportes)
                                if menu_reportes == 1:
                                    reporte_clientes()
                                elif menu_reportes == 2:
                                    reporte_cliente_especifico()
                                elif menu_reportes == 3:
                                    ruta_actual.append('Listado Unidades')
                                    while True:
                                        print('')
                                        mostrar_ruta()
                                        print(f''''
{'=' * 15}MENU DE UNIDADES{'=' * 15}
1.-COMPLETO
2.-POR RODADA 
3.-POR COLOR
4.-VOLVER EL MENU DE INFORMES
    ''')
                                        menu_unidades = input('\nIngresa la opcion deseada: ')
                                        if menu_unidades.isdigit():
                                            menu_unidades = int(menu_unidades)
                                            if menu_unidades == 1:
                                                nombre_archivo = "unidades.csv"
                                                unidades_df = cargar_unidades()
                                                if unidades_df is not None:
                                                    print(f'{'=' * 35}')
                                                    print(f'{'LISTADO DE UNIDADES COMPLETO':^35}')
                                                    print(f'{'=' * 35}')
                                                    print(unidades_df.to_string(index=False))
                                                    print(f'{'=' * 35}')
                                                    while True:
                                                        opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
                                                        if not opcion_exportar:
                                                            break
                                                        elif opcion_exportar.upper() == 'CSV':
                                                            unidades_df.to_csv('Unidades Completas.csv', index = False )
                                                            break
                                                        elif opcion_exportar.upper() == 'MSEXCEL':
                                                            unidades_df.to_excel('Unidades Completas.xlsx', index = False)
                                                            break
                                                        else:
                                                            print ('\nOpcion no valida')
                                                            continue
                                            elif menu_unidades == 2:
                                                while True:
                                                    rodada = input("¿Cuál rodada buscas en especifico? (20, 26,29): ")
                                                    if rodada.isdigit():
                                                        rodada = int(rodada)
                                                        if rodada == 20 or rodada == 26 or rodada == 29:
                                                            nombre_archivo = "unidades.csv"
                                                            unidades_df = cargar_unidades()
                                                            resultado = listado_por_rodada(unidades_df, rodada)
                                                            print(f'{'=' * 29}')
                                                            print(f'{'LISTADO DE UNIDADES POR RODADA':^29}')
                                                            print(f'{'=' * 29}')
                                                            print(resultado.to_string(index=False))
                                                            print(f'{'=' * 29}')
                                                            break
                                                        else:
                                                            print('Esa rodada no existe. Intente de nuevo.')
                                                            continue
                                                    while True:
                                                            opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
                                                            if not opcion_exportar:
                                                                break
                                                            elif opcion_exportar.upper() == 'CSV':
                                                                unidades_df.to_csv('Unidades Rodada.csv', index = False )
                                                                break
                                                            elif opcion_exportar.upper() == 'MSEXCEL':
                                                                unidades_df.to_excel('Unidades Rodada.xlsx', index = False)
                                                                break
                                                            else:
                                                                print ('\nOpcion no valida')
                                                                continue
                                            elif menu_unidades == 3:
                                                while True:
                                                    nombre_archivo = "unidades.csv"
                                                    unidades_df = cargar_unidades()
                                                    print(f'''\n
{'=' * 12} LISTA DE COLORES {'=' * 12} 
{'=' * 42}
''')
                                                    for color in colores:
                                                        print(f'{color:^42}')
                                                    print(f'{'=' * 42}')
                                                    busqueda_color = input("\n¿Qué color buscas?: ")
                                                    if busqueda_color.upper() in colores:
                                                        resultado = listado_por_color(unidades_df, busqueda_color.upper())
                                                        print(f'{'=' * 29}')
                                                        print(f'{'LISTADO DE UNIDADES POR COLOR':^29}')
                                                        print(f'{'=' * 29}')
                                                        print(resultado.to_string(index=False))
                                                        print(f'{'=' * 29}')
                                                        break
                                                    else:
                                                        print('Ese color no existe.')
                                                while True:
                                                        opcion_exportar = input('\nDesea exportar los datos a csv o MsExcel (si no se desea exportar, dejar vacio): ')
                                                        if not opcion_exportar:
                                                            break
                                                        elif opcion_exportar.upper() == 'CSV':
                                                            unidades_df.to_csv('Unidades Color.csv', index = False )
                                                            break
                                                        elif opcion_exportar.upper() == 'MSEXCEL':
                                                            unidades_df.to_excel('Unidades Color.xlsx', index = False)
                                                            break
                                                        else:
                                                            print ('\nOpcion no valida')
                                                            continue
                                            elif menu_unidades == 4:
                                                break
                                            else:
                                                print('opcion no valida. Vuelva a intentar.')
                                                continue
                                        else:
                                            print('Opcion debe contener solo un numero.')
                                            continue
                                    ruta_actual.pop()
                                elif menu_reportes == 4:
                                    retrasos()
                                elif menu_reportes == 5:
                                    reporte_por_retornar()
                                elif menu_reportes == 6:
                                    reporte_por_periodo()
                                elif menu_reportes == 7:
                                    break
                                else:
                                    print('Opcion no valida.')
                                    continue
                            else:
                                print('Solo se permiten opciones numerica.')
                                continue
                        ruta_actual.pop()
                    elif menu_informes == 2:
                        ruta_actual.append('Analisis')
                        while True:
                            print('')
                            mostrar_ruta()
                            print(f'''\n
{'=' * 15}MENU DE ANALISIS{'=' * 15}
1.-DURACION DE LOS PRESTAMOS
2.-RANKING DE CLIENTES 
3.-PREFERENCIAS DE VENTAS
4.-VOLVER EL MENU DE INFORMES
    ''')
                            menu_analisis = input('Ingrese la opcion deseada:')
                            if menu_analisis.isdigit():
                                menu_analisis = int(menu_analisis)
                                if menu_analisis == 1:
                                    calcular_duracion_prestamos()
                                    pass 
                                elif menu_analisis == 2:
                                    reporte_ranking_clientes()
                                    pass 
                                elif menu_analisis == 3:
                                    ruta_actual.append('Preferencias de ventas')
                                    while True:
                                        print('')
                                        mostrar_ruta()
                                        print(f'''\n
{'=' * 15}OBSERVACION EN GRAFICAS{'=' * 15}
1.-RODADA
2.-COLOR
3.-POR DIA
4.-VOLVER EL MENU DE ANALISIS''')
                                        opciones = input("'\n¿Que opcion desea?: ")
                                        if opciones.isdigit():
                                            opciones = int(opciones)
                                            if opciones == 1:
                                                print(f'{'=' * 29}')
                                                print(f'{'REPORTE POR RODADA':^29}')
                                                print(f'{'=' * 29}')
                                                reporte_rodada = reporte_preferencia_rodada()
                                                print(f'{'=' * 29}')
                                                while True:
                                                    opcion1 = input("\n¿Desea observar la grafica de pastel por rodada? (SI/NO): ")
                                                    if opcion1.isalpha():
                                                        if not opcion1:
                                                            print("\nLa opción elegida no puede estar vacía")
                                                            continue
                                                        elif opcion1.upper() == 'SI':
                                                            graficar_prestamos_rodada(reporte_rodada)
                                                            continue 
                                                        elif opcion1.upper() == 'NO':
                                                            break 
                                                        else:
                                                            print('\nNo se permiten respuestas distintas a SI o NO.')
                                                            continue
                                                    else: 
                                                        print("\nLa respuesta debe ser SI o NO.")
                                                        continue
                                            if opciones == 2:
                                                print(f'{'=' * 29}')
                                                print(f'{'REPORTE POR COLOR':^29}')
                                                print(f'{'=' * 29}')
                                                reporte_color = reporte_preferencia_color()
                                                while True:
                                                    opcion2 = input("\n¿Desea observar la grafica de pastel por color? (SI/NO): ")
                                                    if opcion2.isalpha():
                                                        if not opcion2:
                                                            print("\nLa opción elegida no puede estar vacía.")
                                                            continue
                                                        elif opcion2.upper() == 'SI': 
                                                            graficar_prestamos_color(reporte_color)
                                                            continue 
                                                        elif opcion2.upper() == 'NO':
                                                            break 
                                                        else:
                                                            print('\nNo se permiten respuestas distintas a SI o NO.')
                                                            continue
                                                    else:
                                                        print("\nLa respuesta debe ser SI o NO.")
                                                        continue
                                            if opciones == 3:
                                                print(f'{'=' * 29}')
                                                print(f'{'REPORTE POR DIA DE LA SEMANA':^29}')
                                                print(f'{'=' * 29}')
                                                prestamos_dia = prestamos_por_dia()
                                                print(f'{'=' * 29}')
                                                while True:
                                                    opcion3 = input("\n¿Desea observar la grafica por dia? (SI/NO): ")
                                                    if opcion3.isalpha():
                                                        if not opcion3:
                                                            print("\nLa opción elegida no puede estar vacía")
                                                            continue
                                                        elif opcion3.upper() == 'SI': 
                                                            graficar_prestamos_dia(prestamos_dia) 
                                                            break
                                                        elif opcion3.upper() == 'NO':
                                                            break 
                                                        else:
                                                            print('\nNo se permiten respuestas distintas a SI o NO.')
                                                            continue
                                                    else:
                                                        print("\nLa respuesta debe ser SI o NO.")
                                                        continue
                                            if opciones == 4:
                                                opcion4 = input("\n¿Quieres regresar al menu de analisis? (SI/NO: ")
                                                if opcion4.isalpha():
                                                    if not opcion4:
                                                        print("\nLa opcion elegida no puede estar vacia.")
                                                        continue
                                                    elif opcion4.upper() == 'SI':
                                                        break
                                                    elif opcion4.upper() == 'NO':
                                                        continue
                                                else:
                                                    print("\nLa opción elegida deber ser SI o NO")
                                                    continue
                                    ruta_actual.pop()
                                elif menu_analisis == 4:
                                    break
                            else: 
                                print('Opcion no disponible. Intente de nuevo.')
                        ruta_actual.pop()
                    elif menu_informes == 3:
                        break
                    else:
                        print('Esa opcion no existe. Intente de nuevo.')
                        continue
                else:
                    print('Solo se permiten opciones numericas.')
                    continue
            ruta_actual.pop()
        elif menu_principal == 5:
            while True:
                salir = input('Estas seguro que deseas salir del programa?(SI/NO): ')
                if salir.upper() == 'SI':
                    break
                elif salir.upper() == 'NO':
                    break
                else:
                    print('Opcion no valida.')
                    continue
            if salir.upper() == 'SI':
                break
        else:
            print('Opcion no valida.')
        continue
    else:
        print('Solo se permiten caracteres numericos')
        continue